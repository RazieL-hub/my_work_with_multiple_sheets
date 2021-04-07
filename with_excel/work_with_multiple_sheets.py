import openpyxl

book = openpyxl.open('trekking3.xlsx', read_only=True)
directory = book.worksheets[0]  # Справочник калорий
schedule = book.worksheets[1]  # Расписание: день, что едим, в каких количествах
daily_calories = {} # Словарь, где в дне будет записываться количество потребляемых калорий, белков, жиров и углеводов
for row in range(2, schedule.max_row + 1):
    if schedule[row][0].value not in daily_calories.keys():
        daily_calories[schedule[row][0].value] = []
total_calories = 0
total_proteins = 0
total_fats = 0
total_carbohydrates = 0
for day in daily_calories.keys():
    for row_1 in range(2, schedule.max_row+1):
        if day == schedule[row_1][0].value:
            for row_0 in range(2, directory.max_row+1):
                if directory[row_0][0].value == schedule[row_1][1].value:
                    total_calories += (float(directory[row_0][1].value) / 100 * float(schedule[row_1][2].value))

                    if directory[row_0][2].value is None:
                        total_proteins += (0/ 100 * float(schedule[row_1][2].value))
                    else:
                        total_proteins += (float(directory[row_0][2].value) / 100 * float(schedule[row_1][2].value))

                    if directory[row_0][3].value is None:
                        total_fats += (0 / 100 * float(schedule[row_1][2].value))
                    else:
                        total_fats += (float(directory[row_0][3].value) / 100 * float(schedule[row_1][2].value))

                    if directory[row_0][4].value is None:
                        total_carbohydrates += (0 / 100 * float(schedule[row_1][2].value))
                    else:
                        total_carbohydrates += (directory[row_0][4].value / 100 * float(schedule[row_1][2].value))
    daily_calories[day] = [int(total_calories), int(total_proteins), int(total_fats), int(total_carbohydrates)]
    total_calories = 0
    total_proteins = 0
    total_fats = 0
    total_carbohydrates = 0
for i in daily_calories:
    print("В {} день мы съели {} ккал.. Из них белки составляют - {} гр., жиры - {} гр., углеводы - {}гр.".format
          (i, daily_calories[i][0], daily_calories[i][1], daily_calories[i][2], daily_calories[i][3]))